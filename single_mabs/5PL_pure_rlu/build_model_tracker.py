from __future__ import annotations

from pathlib import Path

import openpyxl

SOURCE_XLSX = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/4PL_pure_rlu/model_tracker.xlsx")
OUTPUT_XLSX = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/5PL_pure_rlu/model_tracker.xlsx")

# Row layout of the source tracker (Sheet1), one column per model.
SOURCE_ROWS = {
    "L_random": 3, "L_mab_virus": 4, "L_goes_down": 5,
    "m_random": 6, "m_mab_virus": 7, "m_goes_down": 8,
    "e_random": 9, "e_mab_virus": 10, "e_goes_down": 11,
    "alpha_random": 12, "alpha_run_id": 13,
    "k_random": 14, "k_run_id": 15,
}
SOURCE_MODELS = [f"m{i}" for i in range(64)]  # m0-m63

# Row layout of the new tracker: same as source but with an 's' block
# inserted after 'e' (5PL adds an asymmetry parameter over the 4PL).
ROWS = {
    "L_random": 3, "L_mab_virus": 4, "L_goes_down": 5,
    "m_random": 6, "m_mab_virus": 7, "m_goes_down": 8,
    "e_random": 9, "e_mab_virus": 10, "e_goes_down": 11,
    "s_random": 12, "s_mab_virus": 13, "s_goes_down": 14,
    "alpha_random": 15, "alpha_run_id": 16,
    "k_random": 17, "k_run_id": 18,
}
PARAM_LABELS = {
    3: "L", 6: "m", 9: "e", 12: "s", 15: "alpha", 17: "k",
}
EFFECT_LABELS = {
    3: "Random effect", 4: "mab_virus fixed effect", 5: "goes_down fixed effect",
    6: "Random effect", 7: "mab_virus fixed effect", 8: "goes_down fixed effect",
    9: "Random effect", 10: "mab_virus fixed effect", 11: "goes_down fixed effect",
    12: "Random effect", 13: "mab_virus fixed effect", 14: "goes_down fixed effect",
    15: "Random effect", 16: "run_id fixed effect",
    17: "Random effect", 18: "run_id fixed effect",
}

YES, NO = "Yes", "No"

# (s_random, s_mab_virus) states; s_goes_down is always Yes.
S_STATES = [(NO, NO), (NO, YES), (YES, NO), (YES, YES)]


def read_source_configs(path: Path) -> dict[str, dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    model_cols = {name: idx for idx, name in enumerate(header, start=1) if name in SOURCE_MODELS}
    missing = [m for m in SOURCE_MODELS if m not in model_cols]
    if missing:
        raise SystemExit(f"source tracker is missing model(s): {missing}")

    configs = {}
    for model_name, col in model_cols.items():
        configs[model_name] = {
            field: ws.cell(row=row, column=col).value for field, row in SOURCE_ROWS.items()
        }
    return configs


def main() -> None:
    configs = read_source_configs(SOURCE_XLSX)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.cell(row=2, column=1, value="Parameter")
    for row, label in PARAM_LABELS.items():
        ws.cell(row=row, column=1, value=label)
    for row, label in EFFECT_LABELS.items():
        ws.cell(row=row, column=2, value=label)

    next_col = 3
    next_index = 0
    for source_model in SOURCE_MODELS:
        src = configs[source_model]
        for s_random, s_mab in S_STATES:
            col = next_col
            model_name = f"m{next_index}"

            ws.cell(row=1, column=col, value=model_name)

            ws.cell(row=ROWS["L_random"], column=col, value=src["L_random"])
            ws.cell(row=ROWS["L_mab_virus"], column=col, value=src["L_mab_virus"])
            ws.cell(row=ROWS["L_goes_down"], column=col, value=src["L_goes_down"])

            ws.cell(row=ROWS["m_random"], column=col, value=src["m_random"])
            ws.cell(row=ROWS["m_mab_virus"], column=col, value=src["m_mab_virus"])
            ws.cell(row=ROWS["m_goes_down"], column=col, value=src["m_goes_down"])

            ws.cell(row=ROWS["e_random"], column=col, value=src["e_random"])
            ws.cell(row=ROWS["e_mab_virus"], column=col, value=src["e_mab_virus"])
            ws.cell(row=ROWS["e_goes_down"], column=col, value=src["e_goes_down"])

            ws.cell(row=ROWS["s_random"], column=col, value=s_random)
            ws.cell(row=ROWS["s_mab_virus"], column=col, value=s_mab)
            ws.cell(row=ROWS["s_goes_down"], column=col, value=YES)

            ws.cell(row=ROWS["alpha_random"], column=col, value=src["alpha_random"])
            ws.cell(row=ROWS["alpha_run_id"], column=col, value=src["alpha_run_id"])
            ws.cell(row=ROWS["k_random"], column=col, value=src["k_random"])
            ws.cell(row=ROWS["k_run_id"], column=col, value=src["k_run_id"])

            next_index += 1
            next_col += 1

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"wrote {OUTPUT_XLSX} with {next_index} models (m0-m{next_index - 1})")


if __name__ == "__main__":
    main()
