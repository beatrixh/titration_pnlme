from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
import pandas as pd

BASE_DIR = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs")


def _discover_projects() -> dict[str, Path]:
    """Every subdirectory of BASE_DIR with a model_tracker.xlsx is a known
    project, keyed by its directory name -- so new ones (e.g. a future
    4PL_pure_rlu_v3) show up automatically with no edits needed here.
    "4PL"/"5PL" are kept as short aliases for the two names people actually
    type instead of the full "*_edge_effects" directory name.

    Some projects (e.g. 4PL_plate_fit) nest their versions one level deeper
    (4PL_plate_fit/v2, /v3, ...) instead of the flat *_v2/*_v3 sibling-directory
    convention used elsewhere -- those are discovered too and keyed as
    "{parent}_{child}" (e.g. "4PL_plate_fit_v2") to match that convention.
    """
    projects = {
        p.name: p for p in BASE_DIR.iterdir()
        if p.is_dir() and (p / "model_tracker.xlsx").exists()
    }
    for parent in BASE_DIR.iterdir():
        if not parent.is_dir():
            continue
        for child in parent.iterdir():
            if child.is_dir() and (child / "model_tracker.xlsx").exists():
                projects[f"{parent.name}_{child.name}"] = child
    if "4PL_edge_effects" in projects:
        projects["4PL"] = projects["4PL_edge_effects"]
    if "5PL_edge_effects" in projects:
        projects["5PL"] = projects["5PL_edge_effects"]
    return projects


PROJECTS = _discover_projects()

EFFECT_LABEL = {
    "Random effect": "random",
    "mab_virus fixed effect": "mab_virus",
    "goes_down fixed effect": "goes_down",
    "run_id fixed effect": "run_id",
}


def read_model_tracker(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header = [c.value for c in ws[1]]
    model_cols = {
        idx: name for idx, name in enumerate(header, start=1)
        if name and idx > 2
    }

    rows: dict[str, dict[str, str]] = {name: {} for name in model_cols.values()}
    current_param = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value:
            current_param = row[0].value
        effect = row[1].value
        if effect is None or current_param is None:
            continue
        toggle_name = f"{current_param}_{EFFECT_LABEL.get(effect, effect)}"
        for col_idx, model_name in model_cols.items():
            rows[model_name][toggle_name] = row[col_idx - 1].value

    df = pd.DataFrame.from_dict(rows, orient="index")
    df.index.name = "model"
    return df.reset_index()


def discover_models(models_dir: Path) -> list[str]:
    names = [
        d.name for d in models_dir.iterdir()
        if d.is_dir() and (d / "LogLikelihood" / "logLikelihood.txt").exists()
    ]
    return sorted(names, key=lambda n: int(n[1:]) if n[1:].isdigit() else 10**9)


def read_criteria(model_name: str, models_dir: Path) -> dict:
    path = models_dir / model_name / "LogLikelihood" / "logLikelihood.txt"
    if not path.exists():
        return {"model": model_name, "AIC": None, "BICc": None}
    values = pd.read_csv(path).set_index("criteria")["importanceSampling"]
    return {
        "model": model_name,
        "AIC": values.get("AIC"),
        "BICc": values.get("BICc"),
    }


def build_report(project: str) -> pd.DataFrame:
    root = PROJECTS[project]
    models_dir = root / "model_files"
    tracker_path = root / "model_tracker.xls
x"
    model_names = discover_models(models_dir)
    if not model_names:
        raise SystemExit(f"no completed models (with LogLikelihood/logLikelihood.txt) found in {models_dir}")

    toggles = read_model_tracker(tracker_path)
    criteria = pd.DataFrame([read_criteria(m, models_dir) for m in model_names])

    missing = sorted(set(model_names) - set(toggles["model"]))
    if missing:
        print(f"warning: model(s) not found in tracker: {', '.join(missing)}")

    result = criteria.merge(toggles, on="model", how="left")
    toggle_cols = [c for c in toggles.columns if c != "model"]
    result = result[["model"] + toggle_cols + ["AIC", "BICc"]]
    return result.sort_values("BICc", na_position="last").reset_index(drop=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Report AIC/BICc across all completed models for a project.")
    parser.add_argument("project", choices=sorted(PROJECTS), help="which project to report on")
    args = parser.parse_args()

    result = build_report(args.project)
    print(result.to_string(index=False))

    out_path = PROJECTS[args.project] / f"{args.project.lower()}_likelihood_report.csv"
    result.to_csv(out_path, index=False)
    print(f"\nwrote {out_path}")


if __name__ == "__main__":
    main()
