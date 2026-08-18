from __future__ import annotations

import itertools
from pathlib import Path

import openpyxl

TRACKER_XLSX = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/5PL_edge_effects/model_tracker.xlsx")

# Row layout in Sheet1, fixed by the existing m0 column.
ROWS = {
    "U_random": 3, "U_mab_virus": 4, "U_goes_down": 5,
    "L_random": 6, "L_mab_virus": 7, "L_goes_down": 8,
    "m_random": 9, "m_mab_virus": 10, "m_goes_down": 11,
    "e_random": 12, "e_mab_virus": 13, "e_goes_down": 14,
    "s_random": 15, "s_mab_virus": 16, "s_goes_down": 17,
    "alpha_random": 18, "alpha_run_id": 19,
    "k_random": 20, "k_run_id": 21,
}

YES, NO = "Yes", "No"

# (random_effect, mab_virus_effect) states; goes_down is locked on for
# L, m, e, s throughout and is not part of the factorial.
STATES = [(NO, NO), (NO, YES), (YES, NO), (YES, YES)]

# Valid (m, e) state pairs: every combination of STATES x STATES except the
# one where both m's and e's random effect are off simultaneously.
ME_PAIRS = [
    (m_state, e_state)
    for m_state, e_state in itertools.product(STATES, STATES)
    if m_state[0] == YES or e_state[0] == YES
]


def main() -> None:
    wb = openpyxl.load_workbook(TRACKER_XLSX)
    ws = wb.active

    header = [c.value for c in ws[1]]
    existing_indices = [
        int(v[1:]) for v in header if isinstance(v, str) and v.startswith("m") and v[1:].isdigit()
    ]
    next_index = max(existing_indices, default=-1) + 1
    next_col = ws.max_column + 1

    combos = itertools.product(STATES, STATES, ME_PAIRS)  # (L, s, (m, e))
    n_written = 0
    for l_state, s_state, (m_state, e_state) in combos:
        l_random, l_mab = l_state
        s_random, s_mab = s_state
        m_random, m_mab = m_state
        e_random, e_mab = e_state

        col = next_col
        model_name = f"m{next_index}"

        ws.cell(row=1, column=col, value=model_name)

        ws.cell(row=ROWS["U_random"], column=col, value=NO)
        ws.cell(row=ROWS["U_mab_virus"], column=col, value=NO)
        ws.cell(row=ROWS["U_goes_down"], column=col, value=NO)

        ws.cell(row=ROWS["L_random"], column=col, value=l_random)
        ws.cell(row=ROWS["L_mab_virus"], column=col, value=l_mab)
        ws.cell(row=ROWS["L_goes_down"], column=col, value=YES)

        ws.cell(row=ROWS["m_random"], column=col, value=m_random)
        ws.cell(row=ROWS["m_mab_virus"], column=col, value=m_mab)
        ws.cell(row=ROWS["m_goes_down"], column=col, value=YES)

        ws.cell(row=ROWS["e_random"], column=col, value=e_random)
        ws.cell(row=ROWS["e_mab_virus"], column=col, value=e_mab)
        ws.cell(row=ROWS["e_goes_down"], column=col, value=YES)

        ws.cell(row=ROWS["s_random"], column=col, value=s_random)
        ws.cell(row=ROWS["s_mab_virus"], column=col, value=s_mab)
        ws.cell(row=ROWS["s_goes_down"], column=col, value=YES)

        ws.cell(row=ROWS["alpha_random"], column=col, value=NO)
        ws.cell(row=ROWS["alpha_run_id"], column=col, value=YES)
        ws.cell(row=ROWS["k_random"], column=col, value=NO)
        ws.cell(row=ROWS["k_run_id"], column=col, value=YES)

        next_index += 1
        next_col += 1
        n_written += 1

    wb.save(TRACKER_XLSX)
    print(f"added {n_written} columns, through m{next_index - 1}")


if __name__ == "__main__":
    main()
