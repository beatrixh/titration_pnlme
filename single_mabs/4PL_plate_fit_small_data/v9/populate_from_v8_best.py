"""Populate the v9 model_tracker from the 10 best v8 models.

For each of the 10 named v8 models, copy its h/f/L_gap/eta_alpha toggles
(tracker rows 3-14) verbatim and pair it with all 2**3 combinations of the
new eta_U toggles (rows 15-17: Random effect, mab_virus, run_id).

Result: 80 models, parent-major (m1-m8 from the first parent, etc.), written
to columns C.. of Sheet1. A sidecar CSV records the provenance.
"""
from __future__ import annotations

import csv
import itertools
from pathlib import Path

import openpyxl

BASE = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/4PL_plate_fit_small_data")
V8_TRACKER = BASE / "v8" / "model_tracker.xlsx"
V9_TRACKER = BASE / "v9" / "model_tracker.xlsx"
PROVENANCE_CSV = BASE / "v9" / "v9_model_provenance.csv"

PARENTS = ["m414", "m407", "m472", "m189", "m424", "m471", "m480", "m165", "m406", "m495"]

COPY_ROWS = list(range(3, 15))   # h/f/L_gap/eta_alpha toggles, copied from the v8 parent
ETA_U_ROWS = [15, 16, 17]        # eta_U: Random effect, mab_virus fixed effect, run_id fixed effect
FIRST_MODEL_COL = 3              # column C

YES, NO = "Yes", "No"


def load_v8_configs() -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(V8_TRACKER, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    col_of = {v: i + 1 for i, v in enumerate(header) if isinstance(v, str)}
    missing = [m for m in PARENTS if m not in col_of]
    if missing:
        raise SystemExit(f"v8 tracker is missing model(s): {missing}")
    return {m: [ws.cell(row=r, column=col_of[m]).value for r in COPY_ROWS] for m in PARENTS}


def main() -> None:
    v8 = load_v8_configs()

    wb = openpyxl.load_workbook(V9_TRACKER)
    ws = wb.active

    # Sanity-check the v9 eta_U rows are where we expect them.
    assert ws.cell(row=15, column=1).value == "eta_U", ws.cell(row=15, column=1).value
    assert ws.cell(row=16, column=2).value == "mab_virus fixed effect"
    assert ws.cell(row=17, column=2).value == "run_id fixed effect"

    if ws.max_column >= FIRST_MODEL_COL:
        ws.delete_cols(FIRST_MODEL_COL, ws.max_column - FIRST_MODEL_COL + 1)

    provenance = []
    col = FIRST_MODEL_COL
    index = 1
    for parent in PARENTS:
        parent_cfg = v8[parent]
        for combo in itertools.product((NO, YES), repeat=3):  # (eta_U random, mab_virus, run_id)
            model = f"m{index}"
            ws.cell(row=1, column=col, value=model)
            for r, val in zip(COPY_ROWS, parent_cfg):
                ws.cell(row=r, column=col, value=val)
            for r, val in zip(ETA_U_ROWS, combo):
                ws.cell(row=r, column=col, value=val)
            provenance.append({
                "v9_model": model,
                "v8_parent": parent,
                "eta_U_random": combo[0],
                "eta_U_mab_virus": combo[1],
                "eta_U_run_id": combo[2],
            })
            index += 1
            col += 1

    wb.save(V9_TRACKER)

    with open(PROVENANCE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(provenance[0]))
        w.writeheader()
        w.writerows(provenance)

    print(f"wrote {index - 1} models (m1..m{index - 1}) to {V9_TRACKER}")
    print(f"wrote provenance for {len(provenance)} models to {PROVENANCE_CSV}")


if __name__ == "__main__":
    main()
