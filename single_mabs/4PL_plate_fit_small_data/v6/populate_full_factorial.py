from __future__ import annotations

import itertools
from pathlib import Path

import openpyxl

TRACKER_XLSX = Path(
    "/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/"
    "4PL_plate_fit_small_data/v6/model_tracker.xlsx"
)

# Sheet1 layout, fixed by the existing m0 column:
#   col 1-2  -> parameter / toggle labels
#   col 3    -> m0 (left untouched)
#   row 3-11 -> nine independent Yes/No toggles, in this order:
TOGGLE_ROWS = [
    3,  # h     Random effect
    4,  # h     mab_virus fixed effect
    5,  # h     run_id fixed effect
    6,  # f     Random effect
    7,  # f     mab_virus fixed effect
    8,  # f     run_id fixed effect
    9,  # L_gap Random effect
    10,  # L_gap mab_virus fixed effect
    11,  # L_gap run_id fixed effect
]

FIRST_MODEL_COL = 4  # m0 is col 3; the factorial starts right after it
YES, NO = "Yes", "No"


def main() -> None:
    wb = openpyxl.load_workbook(TRACKER_XLSX)
    ws = wb.active

    # Rebuild from scratch: drop every column after m0, then write the full
    # 2**9 factorial as m1..m512. Idempotent -- safe to re-run.
    if ws.max_column >= FIRST_MODEL_COL:
        ws.delete_cols(FIRST_MODEL_COL, ws.max_column - FIRST_MODEL_COL + 1)

    col = FIRST_MODEL_COL
    for index, combo in enumerate(itertools.product((NO, YES), repeat=9), start=1):
        ws.cell(row=1, column=col, value=f"m{index}")
        for row, value in zip(TOGGLE_ROWS, combo):
            ws.cell(row=row, column=col, value=value)
        col += 1

    wb.save(TRACKER_XLSX)
    print(f"wrote {col - FIRST_MODEL_COL} models, m1 through m{col - FIRST_MODEL_COL}")


if __name__ == "__main__":
    main()
