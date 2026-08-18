from __future__ import annotations

import itertools
from pathlib import Path

import openpyxl

TRACKER_XLSX = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/4PL_pure_rlu/model_tracker.xlsx")

# Row layout in Sheet1, fixed by the existing m0 column.
ROWS = {
    "L_random": 3, "L_mab_virus": 4, "L_goes_down": 5,
    "m_random": 6, "m_mab_virus": 7, "m_goes_down": 8,
    "e_random": 9, "e_mab_virus": 10, "e_goes_down": 11,
    "alpha_random": 12, "alpha_run_id": 13,
    "k_random": 14, "k_run_id": 15,
}

YES, NO = "Yes", "No"

# (random_effect, mab_virus_effect) states toggled per parameter (L, m, e).
STATES = [(NO, NO), (NO, YES), (YES, NO), (YES, YES)]

# goes_down is applied uniformly across L, m, e: all 64 random/mab_virus
# permutations with goes_down on, then all 64 again with goes_down off.
GOES_DOWN_BLOCKS = [YES, NO]


def main() -> None:
    wb = openpyxl.load_workbook(TRACKER_XLSX)
    ws = wb.active

    header = [c.value for c in ws[1]]
    existing_indices = [
        int(v[1:]) for v in header if isinstance(v, str) and v.startswith("m") and v[1:].isdigit()
    ]
    next_index = max(existing_indices, default=-1) + 1
    start_index = next_index
    next_col = ws.max_column + 1

    for goes_down in GOES_DOWN_BLOCKS:
        for combo in itertools.product(STATES, STATES, STATES):  # (L, m, e)
            (l_random, l_mab), (m_random, m_mab), (e_random, e_mab) = combo

            col = next_col
            model_name = f"m{next_index}"

            ws.cell(row=1, column=col, value=model_name)

            ws.cell(row=ROWS["L_random"], column=col, value=l_random)
            ws.cell(row=ROWS["L_mab_virus"], column=col, value=l_mab)
            ws.cell(row=ROWS["L_goes_down"], column=col, value=goes_down)

            ws.cell(row=ROWS["m_random"], column=col, value=m_random)
            ws.cell(row=ROWS["m_mab_virus"], column=col, value=m_mab)
            ws.cell(row=ROWS["m_goes_down"], column=col, value=goes_down)

            ws.cell(row=ROWS["e_random"], column=col, value=e_random)
            ws.cell(row=ROWS["e_mab_virus"], column=col, value=e_mab)
            ws.cell(row=ROWS["e_goes_down"], column=col, value=goes_down)

            ws.cell(row=ROWS["alpha_random"], column=col, value=NO)
            ws.cell(row=ROWS["alpha_run_id"], column=col, value=YES)
            ws.cell(row=ROWS["k_random"], column=col, value=NO)
            ws.cell(row=ROWS["k_run_id"], column=col, value=YES)

            next_index += 1
            next_col += 1

    wb.save(TRACKER_XLSX)
    print(f"added {next_index - start_index} columns, m{start_index} through m{next_index - 1}")


if __name__ == "__main__":
    main()
