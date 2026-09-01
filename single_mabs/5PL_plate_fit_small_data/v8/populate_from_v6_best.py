"""Populate the 5PL v8 model_tracker from the 50 best v6 models.

Take the 50 lowest-BICc models from the v6 likelihood report, copy each one's
h/f/L_gap/s toggles (tracker rows 3-14) verbatim, and pair it with all 2**3
combinations of the new eta_alpha toggles (rows 15-17: Random effect,
mab_virus, run_id).

Result: 400 models, parent-major (m1-m8 from the best v6 model, etc.), written
to columns C.. of Sheet1. A sidecar CSV records the provenance.
"""
from __future__ import annotations

import csv
import itertools
from pathlib import Path

import openpyxl
import pandas as pd

BASE = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/5PL_plate_fit_small_data")
V6_TRACKER = BASE / "v6" / "model_tracker.xlsx"
V6_REPORT = BASE / "v6" / "5pl_plate_fit_small_data_v6_likelihood_report.csv"
V8_TRACKER = BASE / "v8" / "model_tracker.xlsx"
PROVENANCE_CSV = BASE / "v8" / "v8_model_provenance.csv"

N_PARENTS = 50
COPY_ROWS = list(range(3, 15))   # h/f/L_gap/s toggles, copied from the v6 parent
ETA_ALPHA_ROWS = [15, 16, 17]    # eta_alpha: Random effect, mab_virus fixed effect, run_id fixed effect
FIRST_MODEL_COL = 3              # column C

YES, NO = "Yes", "No"


def top_parents() -> list[str]:
    df = pd.read_csv(V6_REPORT)
    return df.nsmallest(N_PARENTS, "BICc")["model"].tolist()


def load_v6_configs(models: list[str]) -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(V6_TRACKER, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    col_of = {v: i + 1 for i, v in enumerate(header) if isinstance(v, str)}
    missing = [m for m in models if m not in col_of]
    if missing:
        raise SystemExit(f"v6 tracker is missing model(s): {missing}")
    return {m: [ws.cell(row=r, column=col_of[m]).value for r in COPY_ROWS] for m in models}


def main() -> None:
    parents = top_parents()
    v6 = load_v6_configs(parents)

    wb = openpyxl.load_workbook(V8_TRACKER)
    ws = wb.active

    # Sanity-check the v8 eta_alpha rows are where we expect them.
    assert ws.cell(row=15, column=1).value == "eta_alpha", ws.cell(row=15, column=1).value
    assert ws.cell(row=16, column=2).value == "mab_virus fixed effect"
    assert ws.cell(row=17, column=2).value == "run_id fixed effect"

    if ws.max_column >= FIRST_MODEL_COL:
        ws.delete_cols(FIRST_MODEL_COL, ws.max_column - FIRST_MODEL_COL + 1)

    provenance = []
    col = FIRST_MODEL_COL
    index = 1
    for rank, parent in enumerate(parents, start=1):
        parent_cfg = v6[parent]
        for combo in itertools.product((NO, YES), repeat=3):  # (eta_alpha random, mab_virus, run_id)
            model = f"m{index}"
            ws.cell(row=1, column=col, value=model)
            for r, val in zip(COPY_ROWS, parent_cfg):
                ws.cell(row=r, column=col, value=val)
            for r, val in zip(ETA_ALPHA_ROWS, combo):
                ws.cell(row=r, column=col, value=val)
            provenance.append({
                "v8_model": model,
                "v6_parent": parent,
                "v6_bicc_rank": rank,
                "eta_alpha_random": combo[0],
                "eta_alpha_mab_virus": combo[1],
                "eta_alpha_run_id": combo[2],
            })
            index += 1
            col += 1

    wb.save(V8_TRACKER)

    with open(PROVENANCE_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(provenance[0]))
        w.writeheader()
        w.writerows(provenance)

    print(f"wrote {index - 1} models (m1..m{index - 1}) to {V8_TRACKER}")
    print(f"wrote provenance for {len(provenance)} models to {PROVENANCE_CSV}")


if __name__ == "__main__":
    main()
