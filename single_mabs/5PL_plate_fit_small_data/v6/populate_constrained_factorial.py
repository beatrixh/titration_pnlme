from __future__ import annotations

import itertools
from pathlib import Path

import openpyxl

TRACKER_XLSX = Path(
    "/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/"
    "5PL_plate_fit_small_data/v6/model_tracker.xlsx"
)

# Sheet1 layout, fixed by the existing m0 column:
#   col 1-2  -> parameter / toggle labels
#   col 3    -> m0 (left untouched)
#   row 3-14 -> twelve Yes/No toggles: {h, f, L_gap, s} x
#              {Random effect, mab_virus fixed effect, run_id fixed effect}

# Rows held Yes for every generated model (not varied).
FORCED_YES_ROWS = [
    4,   # h     mab_virus fixed effect
    9,   # L_gap Random effect
    11,  # L_gap run_id fixed effect
]

# The 9 toggles that are varied, in the order the factorial iterates them.
FREE_ROWS = [
    3,   # h     Random effect
    5,   # h     run_id fixed effect
    6,   # f     Random effect
    7,   # f     mab_virus fixed effect
    8,   # f     run_id fixed effect
    10,  # L_gap mab_virus fixed effect
    12,  # s     Random effect
    13,  # s     mab_virus fixed effect
    14,  # s     run_id fixed effect
]

ALL_TOGGLE_ROWS = list(range(3, 15))  # 12 rows total
MIN_EFFECTS = 3  # drop models with <3 Yes toggles (i.e. >=10 No across the 12)

FIRST_MODEL_COL = 4  # m0 is col 3; the factorial starts right after it
YES, NO = "Yes", "No"


def main() -> None:
    wb = openpyxl.load_workbook(TRACKER_XLSX)
    ws = wb.active

    # Rebuild from scratch: drop every column after m0, then write the
    # constrained factorial. Idempotent -- safe to re-run.
    if ws.max_column >= FIRST_MODEL_COL:
        ws.delete_cols(FIRST_MODEL_COL, ws.max_column - FIRST_MODEL_COL + 1)

    col = FIRST_MODEL_COL
    index = 1
    skipped = 0
    for combo in itertools.product((NO, YES), repeat=len(FREE_ROWS)):
        n_yes = combo.count(YES) + len(FORCED_YES_ROWS)
        if n_yes < MIN_EFFECTS:
            skipped += 1
            continue

        ws.cell(row=1, column=col, value=f"m{index}")
        for row in FORCED_YES_ROWS:
            ws.cell(row=row, column=col, value=YES)
        for row, value in zip(FREE_ROWS, combo):
            ws.cell(row=row, column=col, value=value)

        index += 1
        col += 1

    n_written = index - 1
    ws.cell(row=16, column=1,
            value=f"{n_written} models: 2^{len(FREE_ROWS)} free toggles "
                  f"(h mab_virus + L_gap random + L_gap run_id forced Yes), "
                  f"minus {skipped} with <{MIN_EFFECTS} effects")

    wb.save(TRACKER_XLSX)
    print(f"wrote {n_written} models (m1..m{n_written}), skipped {skipped}")


if __name__ == "__main__":
    main()
