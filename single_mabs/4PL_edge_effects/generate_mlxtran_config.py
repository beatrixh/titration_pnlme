from __future__ import annotations

import os
import re
import sys
from pathlib import Path

import openpyxl

BASE_MLXTRAN = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/4PL_edge_effects/4PL_edge_effects_m0.mlxtran")
MODEL_TRACKER_XLSX = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/4PL_edge_effects/model_tracker.xlsx")
OUTPUT_DIR = Path("/mnt/c/Users/bhaddock/repos/titration_pnlme/single_mabs/4PL_edge_effects/model_files")

# U_pop is held FIXED at 1 regardless of CONFIG, matching this project's
# assay-normalization convention. Set False to let U_pop float like the rest.
FIX_U_POP_AT_1 = True

PARAM_ORDER = ["L", "U", "alpha", "e", "k", "m"]
COVARIATE_ORDER = ["goes_down", "mab_virus", "run_id"]
EFFECT_TO_COVARIATE = {
    "mab_virus fixed effect": "mab_virus",
    "goes_down fixed effect": "goes_down",
    "run_id fixed effect": "run_id",
}


def sanitize(level: str) -> str:
    return "".join("_" if ch in " +-.|" else ch for ch in level)


def beta_name(param: str, covariate: str, level: str) -> str:
    return f"beta_{param}_{covariate}_{sanitize(level)}"


def rewrite_relative_path(text: str, line_pattern: str, base_dir: Path, output_dir: Path) -> str:
    """Re-expresses a relative file path (found via line_pattern's capture
    group) so it's correct from output_dir, instead of from base_dir where
    it was written. Absolute paths are left untouched. This lets the whole
    project directory be moved as a unit and keeps working, since the base
    template and its generated outputs live at different directory depths.
    """
    m = re.search(line_pattern, text)
    if not m:
        return text
    raw_path = m.group(1)
    if re.match(r"^([A-Za-z]:[\\/]|/)", raw_path):
        return text
    abs_target = (base_dir / raw_path).resolve()
    new_rel = os.path.relpath(abs_target, output_dir)
    return text[: m.start(1)] + new_rel + text[m.end(1) :]


def extract_categories(text: str, covariate: str) -> list[str]:
    m = re.search(
        rf"\b{covariate}\s*=\s*\{{type=categorical,\s*categories=\{{(.*?)\}}\}}",
        text,
        re.DOTALL,
    )
    if not m:
        raise ValueError(f"could not find categories for '{covariate}' in base file")
    return re.findall(r"'([^']*)'", m.group(1))


def extract_existing_parameters(text: str) -> dict[str, tuple[str, str]]:
    section = re.search(r"<PARAMETER>(.*?)<MONOLIX>", text, re.DOTALL).group(1)
    out = {}
    for m in re.finditer(r"(\S+)\s*=\s*\{value=([^,]+),\s*method=(\w+)\}", section):
        out[m.group(1)] = (m.group(2), m.group(3))
    return out


def read_model_configs(path: Path) -> dict[str, dict]:
    """Parses model_tracker.xlsx into {model_name: {param: {"variability": bool, "covariates": [...]}}}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header = [c.value for c in ws[1]]
    model_cols = {
        idx: name for idx, name in enumerate(header, start=1)
        if name and idx > 2
    }

    configs: dict[str, dict] = {name: {} for name in model_cols.values()}
    current_param = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value:
            current_param = row[0].value
        effect = row[1].value
        if effect is None or current_param is None:
            continue
        for col_idx, model_name in model_cols.items():
            cfg = configs[model_name].setdefault(
                current_param, {"variability": False, "covariates": []}
            )
            is_yes = str(row[col_idx - 1].value).strip().lower() == "yes"
            if effect == "Random effect":
                cfg["variability"] = is_yes
            elif effect in EFFECT_TO_COVARIATE and is_yes:
                cfg["covariates"].append(EFFECT_TO_COVARIATE[effect])
    return configs


def build_definition_line(param: str, cfg: dict, categories: dict[str, list[str]]) -> str:
    pop = f"{param}_pop"
    covs = cfg["covariates"]
    variability = f"sd=omega_{param}" if cfg["variability"] else "no-variability"

    if not covs:
        return f"{param} = {{distribution=logNormal, typical={pop}, {variability}}}"

    def coeffs(cov: str) -> str:
        betas = [beta_name(param, cov, lvl) for lvl in categories[cov][1:]]
        return "0, " + ", ".join(betas)

    if len(covs) == 1:
        cov = covs[0]
        return (
            f"{param} = {{distribution=logNormal, typical={pop}, covariate={cov}, "
            f"coefficient={{{coeffs(cov)}}}, {variability}}}"
        )

    cov_list = "{" + ", ".join(covs) + "}"
    coeff_blocks = ", ".join("{" + coeffs(c) + "}" for c in covs)
    return (
        f"{param} = {{distribution=logNormal, typical={pop}, covariate={cov_list}, "
        f"coefficient={{{coeff_blocks}}}, {variability}}}"
    )


def build_individual_block(config: dict, categories: dict[str, list[str]]) -> str:
    used_covs = [
        c for c in COVARIATE_ORDER
        if any(c in cfg["covariates"] for cfg in config.values())
    ]

    input_names = []
    for p in PARAM_ORDER:
        input_names.append(f"{p}_pop")
        if config[p]["variability"]:
            input_names.append(f"omega_{p}")
    for cov in used_covs:
        input_names.append(cov)
        for p in PARAM_ORDER:
            if cov in config[p]["covariates"]:
                for lvl in categories[cov][1:]:
                    input_names.append(beta_name(p, cov, lvl))

    lines = ["[INDIVIDUAL]", "input = {" + ", ".join(input_names) + "}", ""]
    for cov in used_covs:
        cats = ", ".join(f"'{lvl}'" for lvl in categories[cov])
        lines.append(f"{cov} = {{type=categorical, categories={{{cats}}}}}")
    if used_covs:
        lines.append("")
    lines.append("DEFINITION:")
    for p in PARAM_ORDER:
        lines.append(build_definition_line(p, config[p], categories))
    return "\n".join(lines)


def build_parameter_block(
    config: dict, categories: dict[str, list[str]], existing: dict[str, tuple[str, str]]
) -> str:
    lines = ["<PARAMETER>"]

    def emit(name: str, default_value: str, default_method: str) -> None:
        value, method = existing.get(name, (default_value, default_method))
        lines.append(f"{name} = {{value={value}, method={method}}}")

    for p in PARAM_ORDER:
        if p == "U" and FIX_U_POP_AT_1:
            lines.append("U_pop = {value=1, method=FIXED}")
        else:
            emit(f"{p}_pop", "1", "MLE")
        if config[p]["variability"]:
            emit(f"omega_{p}", "1", "MLE")
        for cov in config[p]["covariates"]:
            for lvl in categories[cov][1:]:
                emit(beta_name(p, cov, lvl), "0", "MLE")

    for name, default in [("a", ("1", "MLE")), ("b", ("1", "MLE")), ("c", ("1", "FIXED"))]:
        value, method = existing.get(name, default)
        lines.append(f"{name} = {{value={value}, method={method}}}")

    return "\n".join(lines)


def generate_one(model_name: str, config: dict) -> None:
    missing = [p for p in PARAM_ORDER if p not in config]
    if missing:
        raise SystemExit(
            f"model '{model_name}' is missing parameter(s) {missing} in {MODEL_TRACKER_XLSX}"
        )

    output_mlxtran = OUTPUT_DIR / f"4PL_edge_effects_{model_name}.mlxtran"

    with open(BASE_MLXTRAN, "r", newline="") as f:
        text = f.read()
    uses_crlf = "\r\n" in text
    text = text.replace("\r\n", "\n")

    text = rewrite_relative_path(text, r"file=\{path='([^']*)'\}", BASE_MLXTRAN.parent, OUTPUT_DIR)
    text = rewrite_relative_path(text, r"file = '([^']*)'", BASE_MLXTRAN.parent, OUTPUT_DIR)

    categories = {cov: extract_categories(text, cov) for cov in COVARIATE_ORDER}
    existing = extract_existing_parameters(text)

    new_individual = build_individual_block(config, categories)
    new_parameter = build_parameter_block(config, categories, existing)

    individual_start = text.index("[INDIVIDUAL]")
    individual_end = text.index("[LONGITUDINAL]")
    text = text[:individual_start] + new_individual + "\n\n" + text[individual_end:]

    parameter_start = text.index("<PARAMETER>")
    parameter_end = text.index("<MONOLIX>")
    text = text[:parameter_start] + new_parameter + "\n\n" + text[parameter_end:]

    text = re.sub(r"exportpath = '.*?'", f"exportpath = '{model_name}'", text)

    if uses_crlf:
        text = text.replace("\n", "\r\n")
    with open(output_mlxtran, "w", newline="") as f:
        f.write(text)
    print(f"wrote {output_mlxtran}")


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit(f"usage: python3 {Path(sys.argv[0]).name} <model_name|--all>")
    arg = sys.argv[1]

    model_configs = read_model_configs(MODEL_TRACKER_XLSX)

    if arg == "--all":
        model_names = list(model_configs)
    elif arg in model_configs:
        model_names = [arg]
    else:
        raise SystemExit(
            f"'{arg}' not found in {MODEL_TRACKER_XLSX}; available: {sorted(model_configs)}"
        )

    for model_name in model_names:
        generate_one(model_name, model_configs[model_name])


if __name__ == "__main__":
    main()
