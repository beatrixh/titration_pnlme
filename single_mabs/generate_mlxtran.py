from __future__ import annotations

import argparse
import os
import re
from pathlib import Path

import openpyxl

COVARIATE_ORDER = ["goes_down", "mab_virus", "run_id"]
EFFECT_TO_COVARIATE = {
    "mab_virus fixed effect": "mab_virus",
    "goes_down fixed effect": "goes_down",
    "run_id fixed effect": "run_id",
}

# U_pop is held FIXED at 1 regardless of CONFIG, matching this project's
# assay-normalization convention. Pass --no-fix-u to let U_pop float instead.
FIX_U_POP_AT_1 = True


def sanitize(level: str) -> str:
    # Mlxtran identifiers must be ASCII alphanumeric/underscore -- anything
    # else (e.g. the Greek delta in a virus strain name like 'GT1.1ΔCT_His')
    # gets replaced one-for-one, matching how Monolix's own GUI sanitizes
    # category levels into parameter names.
    return re.sub(r"[^A-Za-z0-9_]", "_", level)


def beta_name(param: str, covariate: str, level: str) -> str:
    return f"beta_{param}_{covariate}_{sanitize(level)}"


def rewrite_relative_path(text: str, line_pattern: str, base_dir: Path, output_dir: Path) -> str:
    """Re-expresses a relative file path (found via line_pattern's capture
    group) so it's correct from output_dir, instead of from base_dir where
    it was written. Absolute paths are left untouched. This lets the whole
    project directory be moved as a unit and keeps working, since the
    template and its generated outputs can live at different directory
    depths.
    """
    m = re.search(line_pattern, text)
    if not m:
        return text
    raw_path = m.group(1)
    if re.match(r"^([A-Za-z]:[\\/]|/)", raw_path):
        return text
    abs_target = (base_dir / raw_path).resolve()
    new_rel = os.path.relpath(abs_target, output_dir)
    return text[: m.start(1)] + new_rel + text[m.end(1) :]


def extract_categories(text: str, covariate: str) -> list[str]:
    m = re.search(
        rf"\b{covariate}\s*=\s*\{{type=categorical,\s*categories=\{{(.*?)\}}\}}",
        text,
        re.DOTALL,
    )
    if not m:
        raise ValueError(f"could not find categories for '{covariate}' in template")
    return re.findall(r"'([^']*)'", m.group(1))


def extract_distributions(text: str, params: list[str]) -> dict[str, str]:
    """Reads each parameter's distribution (logNormal, normal, ...) from the
    template's own DEFINITION block, rather than assuming -- some
    parameters use `normal` instead of `logNormal`, presumably to allow
    negative values, and this varies by project.
    """
    section = re.search(r"DEFINITION:(.*?)\[LONGITUDINAL\]", text, re.DOTALL).group(1)
    out = {}
    for param in params:
        m = re.search(rf"^\s*{param}\s*=\s*\{{distribution=(\w+),", section, re.MULTILINE)
        if not m:
            raise ValueError(f"could not find distribution for parameter '{param}' in template")
        out[param] = m.group(1)
    return out


def extract_existing_parameters(text: str) -> dict[str, tuple[str, str]]:
    section = re.search(r"<PARAMETER>(.*?)<MONOLIX>", text, re.DOTALL).group(1)
    out = {}
    for m in re.finditer(r"(\S+)\s*=\s*\{value=([^,]+),\s*method=(\w+)\}", section):
        out[m.group(1)] = (m.group(2), m.group(3))
    return out


def read_model_configs(path: Path) -> tuple[list[str], dict[str, dict]]:
    """Parses model_tracker.xlsx into (param_order, {model_name: {param: {"variability": bool, "covariates": [...]}}}).

    param_order is the list of parameter names (e.g. L, U, alpha, e, k, m,
    s), in the order they first appear down column A -- read from the
    tracker itself rather than hardcoded, so this works for any project's
    parameter set without editing the script.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header = [c.value for c in ws[1]]
    model_cols = {
        idx: name for idx, name in enumerate(header, start=1)
        if name and idx > 2
    }

    param_order: list[str] = []
    configs: dict[str, dict] = {name: {} for name in model_cols.values()}
    current_param = None
    for row in ws.iter_rows(min_row=3):
        if row[0].value:
            current_param = row[0].value
            if current_param not in param_order:
                param_order.append(current_param)
        effect = row[1].value
        if effect is None or current_param is None:
            continue
        for col_idx, model_name in model_cols.items():
            cfg = configs[model_name].setdefault(
                current_param, {"variability": False, "covariates": []}
            )
            is_yes = str(row[col_idx - 1].value).strip().lower() == "yes"
            if effect == "Random effect":
                cfg["variability"] = is_yes
            elif effect in EFFECT_TO_COVARIATE and is_yes:
                cfg["covariates"].append(EFFECT_TO_COVARIATE[effect])
    return param_order, configs


def build_definition_line(
    param: str, cfg: dict, categories: dict[str, list[str]], distribution: str
) -> str:
    pop = f"{param}_pop"
    covs = cfg["covariates"]
    variability = f"sd=omega_{param}" if cfg["variability"] else "no-variability"

    if not covs:
        return f"{param} = {{distribution={distribution}, typical={pop}, {variability}}}"

    def coeffs(cov: str) -> str:
        betas = [beta_name(param, cov, lvl) for lvl in categories[cov][1:]]
        return "0, " + ", ".join(betas)

    if len(covs) == 1:
        cov = covs[0]
        return (
            f"{param} = {{distribution={distribution}, typical={pop}, covariate={cov}, "
            f"coefficient={{{coeffs(cov)}}}, {variability}}}"
        )

    cov_list = "{" + ", ".join(covs) + "}"
    coeff_blocks = ", ".join("{" + coeffs(c) + "}" for c in covs)
    return (
        f"{param} = {{distribution={distribution}, typical={pop}, covariate={cov_list}, "
        f"coefficient={{{coeff_blocks}}}, {variability}}}"
    )


def build_individual_block(
    param_order: list[str],
    config: dict,
    categories: dict[str, list[str]],
    distributions: dict[str, str],
) -> str:
    used_covs = [
        c for c in COVARIATE_ORDER
        if any(c in cfg["covariates"] for cfg in config.values())
    ]

    input_names = []
    for p in param_order:
        input_names.append(f"{p}_pop")
        if config[p]["variability"]:
            input_names.append(f"omega_{p}")
    for cov in used_covs:
        input_names.append(cov)
        for p in param_order:
            if cov in config[p]["covariates"]:
                for lvl in categories[cov][1:]:
                    input_names.append(beta_name(p, cov, lvl))

    lines = ["[INDIVIDUAL]", "input = {" + ", ".join(input_names) + "}", ""]
    for cov in used_covs:
        cats = ", ".join(f"'{lvl}'" for lvl in categories[cov])
        lines.append(f"{cov} = {{type=categorical, categories={{{cats}}}}}")
    if used_covs:
        lines.append("")
    lines.append("DEFINITION:")
    for p in param_order:
        lines.append(build_definition_line(p, config[p], categories, distributions[p]))
    return "\n".join(lines)


def build_parameter_block(
    param_order: list[str],
    config: dict,
    categories: dict[str, list[str]],
    existing: dict[str, tuple[str, str]],
    fix_u_pop_at_1: bool,
) -> str:
    lines = ["<PARAMETER>"]

    def emit(name: str, default_value: str, default_method: str) -> None:
        value, method = existing.get(name, (default_value, default_method))
        lines.append(f"{name} = {{value={value}, method={method}}}")

    for p in param_order:
        if p == "U" and fix_u_pop_at_1:
            lines.append("U_pop = {value=1, method=FIXED}")
        else:
            emit(f"{p}_pop", "1", "MLE")
        if config[p]["variability"]:
            emit(f"omega_{p}", "1", "MLE")
        for cov in config[p]["covariates"]:
            for lvl in categories[cov][1:]:
                emit(beta_name(p, cov, lvl), "0", "MLE")

    for name, default in [("a", ("1", "MLE")), ("b", ("1", "MLE")), ("c", ("1", "FIXED"))]:
        value, method = existing.get(name, default)
        lines.append(f"{name} = {{value={value}, method={method}}}")

    return "\n".join(lines)


def generate_one(
    model_name: str,
    config: dict,
    param_order: list[str],
    template: Path,
    tracker: Path,
    output_dir: Path,
    prefix: str,
    fix_u_pop_at_1: bool,
) -> None:
    missing = [p for p in param_order if p not in config]
    if missing:
        raise SystemExit(f"model '{model_name}' is missing parameter(s) {missing} in {tracker}")

    output_dir.mkdir(parents=True, exist_ok=True)
    output_mlxtran = output_dir / f"{prefix}{model_name}.mlxtran"

    with open(template, "r", newline="") as f:
        text = f.read()
    uses_crlf = "\r\n" in text
    text = text.replace("\r\n", "\n")

    text = rewrite_relative_path(text, r"file=\{path='([^']*)'\}", template.parent, output_dir)
    text = rewrite_relative_path(text, r"file = '([^']*)'", template.parent, output_dir)

    categories = {cov: extract_categories(text, cov) for cov in COVARIATE_ORDER}
    distributions = extract_distributions(text, param_order)
    existing = extract_existing_parameters(text)

    new_individual = build_individual_block(param_order, config, categories, distributions)
    new_parameter = build_parameter_block(param_order, config, categories, existing, fix_u_pop_at_1)

    individual_start = text.index("[INDIVIDUAL]")
    individual_end = text.index("[LONGITUDINAL]")
    text = text[:individual_start] + new_individual + "\n\n" + text[individual_end:]

    parameter_start = text.index("<PARAMETER>")
    parameter_end = text.index("<MONOLIX>")
    text = text[:parameter_start] + new_parameter + "\n\n" + text[parameter_end:]

    text = re.sub(r"exportpath = '.*?'", f"exportpath = '{model_name}'", text)

    if uses_crlf:
        text = text.replace("\n", "\r\n")
    with open(output_mlxtran, "w", newline="") as f:
        f.write(text)
    print(f"wrote {output_mlxtran}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate per-model .mlxtran files from a template + model_tracker.xlsx."
    )
    parser.add_argument("--template", required=True, type=Path, help="base .mlxtran to derive from")
    parser.add_argument("--tracker", required=True, type=Path, help="model_tracker.xlsx path")
    parser.add_argument("--output-dir", required=True, type=Path, help="where generated .mlxtran files go")
    parser.add_argument(
        "--prefix", default="",
        help="filename prefix for generated files, e.g. '5PL_edge_effects_' -> "
             "'5PL_edge_effects_m0.mlxtran'. Defaults to no prefix ('m0.mlxtran').",
    )
    parser.add_argument("--no-fix-u", action="store_true", help="let U_pop float instead of FIXED=1")
    parser.add_argument(
        "models", nargs="+",
        help="model names to generate (e.g. m0 m14 m192), or '--all' for every model in the tracker",
    )
    args = parser.parse_args()

    param_order, model_configs = read_model_configs(args.tracker)

    if args.models == ["--all"]:
        model_names = list(model_configs)
    else:
        unknown = [m for m in args.models if m not in model_configs]
        if unknown:
            raise SystemExit(
                f"model(s) not found in {args.tracker}: {unknown}; "
                f"available: {sorted(model_configs)}"
            )
        model_names = args.models

    for model_name in model_names:
        generate_one(
            model_name,
            model_configs[model_name],
            param_order,
            args.template,
            args.tracker,
            args.output_dir,
            args.prefix,
            fix_u_pop_at_1=not args.no_fix_u,
        )


if __name__ == "__main__":
    main()
